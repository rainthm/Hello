# --*-- coding:utf8 --*--

#-----------------------------------

#Python操作Excel之根据一个工作簿中的内容修改另一个工作簿


import openpyxl
import os


wb1 = openpyxl.load_workbook(r'./leili.xlsx')
wb2 = openpyxl.load_workbook(r'./wait_del.xlsx')
sh1 = wb1['xinqu']
sh2 = wb2.active
name_list = []

col_list = [ col for col in sh2.columns]
for cell in col_list[1]:
    if cell.value != "姓名":
        name_list.append(cell.value)

for i in name_list:
    print(i, end="   ")
print()

for i in range(2, sh1.max_row+1):
    if sh1[f"A{i}"].value in name_list:
        print(f"离职人员:", sh1[f"A{i}"].value)
        sh1[f"C{i}"] = "离职人员"
        print(f"C{i} value: ", sh1[f"C{i}"].value)

wb1.save(r'./leili.xlsx')
